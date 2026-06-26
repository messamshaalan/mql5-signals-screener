#!/usr/bin/env python3
"""
MT5 Signals Screener
Scrapes all MT5 signal providers from MQL5 (12 pages, sorted by gain)
and generates an Excel + interactive HTML report.

Usage:
    python scraper.py          # generate files in ./output/
    python server.py           # run interactive server at localhost:5000
"""

from curl_cffi import requests
from bs4 import BeautifulSoup
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter
import time
import re
import json
import os
import sys
from datetime import datetime

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

# ── Config ─────────────────────────────────────────────────────────────────────

TOTAL_PAGES = 12
BASE_URL    = "https://www.mql5.com"
LIST_URL    = "https://www.mql5.com/en/signals/mt5/list"
DELAY       = 2.5

HEADERS = {
    "Accept-Language": "en-US,en;q=0.9",
    "Referer":         "https://www.mql5.com/en/signals",
}

# ── Helpers ────────────────────────────────────────────────────────────────────

def page_url(n: int) -> str:
    return LIST_URL + ("?orderby=gain" if n == 1 else f"/page{n}?orderby=gain")


def to_float(text: str):
    if not text:
        return None
    text = str(text).strip()
    m = re.match(r"^1:(\d+)$", text)
    if m:
        return float(m.group(1))
    cleaned = re.sub(r"[^\d.KkMm\-]", "", text.replace(",", ""))
    mul = 1
    if cleaned.upper().endswith("K"):
        mul, cleaned = 1_000, cleaned[:-1]
    elif cleaned.upper().endswith("M"):
        mul, cleaned = 1_000_000, cleaned[:-1]
    try:
        return float(cleaned) * mul
    except (ValueError, TypeError):
        return None


def fetch(session, url: str, retries: int = 3) -> str | None:
    for attempt in range(retries):
        try:
            r = session.get(url, headers=HEADERS, timeout=30)
            r.raise_for_status()
            return r.text
        except Exception as e:
            wait = 6 * (attempt + 1)
            print(f"    ✗ attempt {attempt+1}/{retries} — {e}  (retry in {wait}s)")
            if attempt < retries - 1:
                time.sleep(wait)
    return None

# ── Parser ─────────────────────────────────────────────────────────────────────

def _col(row_div, css_class: str) -> str:
    el = row_div.find("div", class_=css_class)
    return el.get_text(separator=" ", strip=True) if el else ""


def parse_page(html: str, page_num: int) -> list[dict]:
    soup = BeautifulSoup(html, "html.parser")
    signal_divs = [
        d for d in soup.find_all("div")
        if d.get("class") and "signal" in d["class"] and "row" in d["class"]
    ]
    if not signal_divs:
        print(f"    ✗ no signal divs found on page {page_num}")
        return []

    signals = []
    for div in signal_divs:
        avatar    = div.find("a", class_="signal-avatar")
        name_span = div.find("span", class_="name")
        name   = name_span.get_text(strip=True) if name_span else ""
        link   = ""
        author = ""
        if avatar:
            href = avatar.get("href", "")
            link = (BASE_URL + href if href.startswith("/") else href).split("?")[0]
            title = avatar.get("title", "")
            m = re.search(r"by (.+)$", title)
            if m:
                author = m.group(1).strip()

        btn       = div.find("div", class_="button")
        signal_id = btn.get("data-id", "") if btn else ""

        gain_raw = _col(div, "col-growth")
        dd_raw   = _col(div, "col-drawdown")
        win_raw  = _col(div, "col-plus")
        pf_raw   = _col(div, "col-pf")

        record = {
            "Rank":             _col(div, "col-num"),
            "Signal ID":        signal_id,
            "Name":             name,
            "Author":           author,
            "Link":             link,
            "Price (USD/mo)":   _col(div, "col-price"),
            "Gain %":           gain_raw,
            "Gain_num":         to_float(gain_raw),
            "Drawdown %":       dd_raw,
            "Drawdown_num":     to_float(dd_raw),
            "Subscribers":      _col(div, "col-subscribers"),
            "Subscriber Funds": _col(div, "col-facilities"),
            "Balance":          _col(div, "col-balance"),
            "Weeks":            _col(div, "col-weeks"),
            "EA %":             _col(div, "col-experts"),
            "Trades":           _col(div, "col-trades"),
            "Win %":            win_raw,
            "Win_num":          to_float(win_raw),
            "Activity %":       _col(div, "col-activity"),
            "Profit Factor":    pf_raw,
            "PF_num":           to_float(pf_raw),
            "Expected Payoff":  _col(div, "col-ep"),
            "Leverage":         _col(div, "col-leverage"),
            "Page":             page_num,
        }
        if record["Name"]:
            signals.append(record)

    return signals

# ── Scraper ────────────────────────────────────────────────────────────────────

def warm_up_session(session):
    try:
        r = session.get("https://www.mql5.com/en", timeout=20)
        print(f"  Session warmup: {r.status_code}  ({len(session.cookies)} cookies)")
        time.sleep(2)
    except Exception as e:
        print(f"  Session warmup skipped: {e}")


def scrape_all() -> list[dict]:
    session     = requests.Session(impersonate="chrome120")
    all_signals: list[dict] = []

    print("  Warming up session (collecting cookies)...")
    warm_up_session(session)

    for page in range(1, TOTAL_PAGES + 1):
        url = page_url(page)
        print(f"  [{page:2d}/{TOTAL_PAGES}] {url}")
        html = fetch(session, url)
        if html:
            sigs = parse_page(html, page)
            all_signals.extend(sigs)
            print(f"         → {len(sigs)} signals  (running total: {len(all_signals)})")
        else:
            print(f"         → FAILED — skipping page {page}")
        if page < TOTAL_PAGES:
            time.sleep(DELAY)

    return all_signals

# ── Excel output ───────────────────────────────────────────────────────────────

DISPLAY_COLS = [
    "Rank", "Signal ID", "Name", "Author", "Price (USD/mo)", "Gain %", "Drawdown %",
    "Win %", "Profit Factor", "Expected Payoff", "Subscribers",
    "Subscriber Funds", "Balance", "Weeks", "Trades", "EA %",
    "Activity %", "Leverage", "Link",
]

COL_WIDTHS = {
    "Rank": 6, "Signal ID": 10, "Name": 32, "Author": 20, "Price (USD/mo)": 14,
    "Gain %": 14, "Drawdown %": 12, "Win %": 9, "Profit Factor": 13,
    "Expected Payoff": 16, "Subscribers": 13, "Subscriber Funds": 16,
    "Balance": 14, "Weeks": 8, "Trades": 8, "EA %": 7,
    "Activity %": 11, "Leverage": 10, "Link": 10,
}


def build_excel(df: pd.DataFrame, path: str):
    cols = [c for c in DISPLAY_COLS if c in df.columns]
    data = df[cols].copy()

    wb = Workbook()
    ws = wb.active
    ws.title = "MT5 Signals"

    HDR_FILL  = PatternFill("solid", fgColor="1F4E79")
    HDR_FONT  = Font(color="FFFFFF", bold=True, size=10)
    HDR_ALIGN = Alignment(horizontal="center", vertical="center", wrap_text=True)
    EVEN_FILL = PatternFill("solid", fgColor="EFF6FF")
    LINK_FONT = Font(color="0563C1", underline="single", size=10)
    CENTER    = Alignment(horizontal="center", vertical="center")
    LEFT      = Alignment(horizontal="left",   vertical="center")

    for ci, name in enumerate(cols, 1):
        c = ws.cell(row=1, column=ci, value=name)
        c.fill, c.font, c.alignment = HDR_FILL, HDR_FONT, HDR_ALIGN
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(name, 12)
    ws.row_dimensions[1].height = 32

    for ri, row in enumerate(data.itertuples(index=False), 2):
        even = ri % 2 == 0
        for ci, val in enumerate(row, 1):
            col_name = cols[ci - 1]
            c = ws.cell(row=ri, column=ci, value=(val if val not in (None, "") else None))
            c.alignment = LEFT if col_name in ("Name", "Author") else CENTER
            if col_name == "Link" and val:
                c.value, c.hyperlink, c.font = "→ View", val, LINK_FONT
                c.alignment = CENTER
            elif even:
                c.fill = EVEN_FILL

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{len(data)+1}"

    n = len(data) + 1
    for col_name, inverted in [("Gain %", False), ("Drawdown %", True), ("Win %", False)]:
        if col_name not in cols:
            continue
        ci  = cols.index(col_name) + 1
        ref = f"{get_column_letter(ci)}2:{get_column_letter(ci)}{n}"
        lo, hi = ("FF6B6B", "57C965") if not inverted else ("57C965", "FF6B6B")
        ws.conditional_formatting.add(ref, ColorScaleRule(
            start_type="min",      start_color=lo,
            mid_type="percentile", mid_value=50, mid_color="FFE066",
            end_type="max",        end_color=hi,
        ))

    wb.save(path)
    print(f"  ✓ Excel  → {path}")


# ── HTML output ────────────────────────────────────────────────────────────────

HTML_DISPLAY = [
    "Rank", "Signal ID", "Name", "Author", "Price (USD/mo)", "Gain %", "Drawdown %",
    "Win %", "Profit Factor", "Expected Payoff", "Subscribers",
    "Subscriber Funds", "Balance", "Weeks", "Trades", "EA %",
    "Activity %", "Leverage",
]


def build_html(df: pd.DataFrame, path: str, scraped_at: str):
    cols = [c for c in HTML_DISPLAY if c in df.columns]

    rows = []
    for _, r in df.iterrows():
        row          = {c: str(r.get(c, "") or "") for c in cols}
        row["__link"] = str(r.get("Link", "") or "")
        row["__id"]   = str(r.get("Signal ID", "") or "")
        rows.append(row)

    rows_json = json.dumps(rows, ensure_ascii=False)
    cols_json = json.dumps(cols)
    total     = len(rows)

    template = r"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1.0">
<title>MT5 Signals Screener</title>
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/css/bootstrap.min.css">
<link rel="stylesheet" href="https://cdn.datatables.net/1.13.8/css/dataTables.bootstrap5.min.css">
<link rel="stylesheet" href="https://cdn.datatables.net/buttons/2.4.2/css/buttons.bootstrap5.min.css">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/nouislider@15.8.1/dist/nouislider.min.css">
<style>
:root{--bg0:#0d1117;--bg1:#161b22;--bg2:#1c2230;--bdr:#30363d;--blue:#1F4E79;--blue2:#2a6099;
  --txt:#c9d1d9;--txt2:#8b949e;--green:#3fb950;--red:#f85149;--yellow:#d29922;--purple:#bc8cff;}
*{box-sizing:border-box;}
body{background:var(--bg0);color:var(--txt);font-family:'Segoe UI',system-ui,sans-serif;min-height:100vh;font-size:14px;}

/* TOPBAR */
.topbar{background:linear-gradient(135deg,#1a3a5c,#0d1f35);padding:.65rem 1.5rem;display:flex;
  align-items:center;justify-content:space-between;box-shadow:0 2px 12px #0009;
  border-bottom:1px solid var(--bdr);position:sticky;top:0;z-index:100;}
.topbar h1{font-size:1.1rem;font-weight:700;margin:0;color:#fff;letter-spacing:.3px;}
.topbar-badge{background:var(--blue);color:#fff;font-size:.62rem;font-weight:700;
  padding:2px 7px;border-radius:10px;letter-spacing:.5px;margin-left:10px;}
.topbar-right{display:flex;align-items:center;gap:12px;}
.topbar-meta{color:var(--txt2);font-size:.74rem;}
.btn-update{background:var(--blue);color:#fff;border:1px solid var(--blue2);border-radius:6px;
  padding:6px 16px;font-size:.8rem;font-weight:600;cursor:pointer;transition:background .15s;
  display:flex;align-items:center;gap:6px;white-space:nowrap;}
.btn-update:hover:not(:disabled){background:var(--blue2);}
.btn-update:disabled{opacity:.6;cursor:not-allowed;}

/* LAYOUT */
.main-wrap{padding:1rem 1.25rem;max-width:1900px;margin:0 auto;}

/* SECTION LABELS */
.section-lbl{font-size:.68rem;font-weight:700;letter-spacing:1.8px;text-transform:uppercase;
  color:var(--txt2);margin:1.25rem 0 .6rem;display:flex;align-items:center;gap:8px;}
.section-lbl::after{content:'';flex:1;height:1px;background:var(--bdr);}

/* TOP PICKS */
.picks-wrap{display:flex;flex-wrap:wrap;gap:8px;margin-bottom:.75rem;}
.pick-card{flex:1 1 185px;max-width:230px;background:var(--bg1);border:1px solid var(--bdr);
  border-radius:10px;padding:11px 13px;cursor:pointer;transition:border-color .2s,box-shadow .2s;}
.pick-card:hover,.pick-card.active{border-color:var(--blue2);box-shadow:0 0 0 1px var(--blue2) inset;}
.pick-rank-row{display:flex;justify-content:space-between;align-items:center;margin-bottom:5px;}
.pick-rank{font-size:.62rem;font-weight:700;color:var(--txt2);background:var(--bg2);
  padding:1px 6px;border-radius:10px;}
.pick-score{font-size:.62rem;font-weight:700;padding:1px 6px;border-radius:10px;}
.pick-name{font-size:.8rem;font-weight:600;color:#58a6ff;text-decoration:none;display:block;
  line-height:1.3;margin-bottom:6px;white-space:nowrap;overflow:hidden;text-overflow:ellipsis;}
.pick-bar-track{height:3px;background:var(--bg2);border-radius:2px;margin-bottom:7px;}
.pick-bar-fill{height:3px;border-radius:2px;}
.pick-stats{display:grid;grid-template-columns:1fr 1fr;gap:2px 8px;font-size:.68rem;}
.pick-stat-lbl{color:var(--txt2);}
.pick-stat-val{font-weight:600;color:var(--txt);text-align:right;}
.pick-badges{margin-top:6px;display:flex;flex-wrap:wrap;gap:3px;}
.pick-badge{font-size:.6rem;background:#1a2a3a;color:#58a6ff;padding:1px 5px;border-radius:8px;border:1px solid #1e3a55;}

/* STAT CARDS */
.stat-card{background:var(--bg1);border:1px solid var(--bdr);border-radius:8px;
  padding:12px 16px;text-align:center;}
.stat-val{font-size:1.3rem;font-weight:700;line-height:1.2;}
.stat-lbl{font-size:.64rem;color:var(--txt2);text-transform:uppercase;letter-spacing:.8px;margin-top:3px;}

/* CHARTS */
.chart-card{background:var(--bg1);border:1px solid var(--bdr);border-radius:10px;
  padding:1rem;height:100%;}
.chart-hdr{font-size:.76rem;font-weight:600;color:var(--txt2);margin-bottom:.75rem;}

/* FILTERS */
.filter-box{background:var(--bg1);border:1px solid var(--bdr);border-radius:10px;
  padding:1rem 1.5rem;margin-bottom:1rem;}
.filter-box label{font-size:.72rem;color:var(--txt2);display:block;margin-bottom:2px;}
.filter-box input[type=range]{width:100%;accent-color:var(--blue2);}
.filter-box input[type=number]{background:var(--bg0);border:1px solid var(--bdr);color:var(--txt);
  border-radius:5px;padding:3px 8px;width:85px;font-size:.77rem;}
.range-val{color:#58a6ff;font-weight:600;font-size:.82rem;}
.preset-btns{display:flex;flex-wrap:wrap;gap:3px;margin-top:4px;}
.preset-btn{background:var(--bg2);border:1px solid var(--bdr);color:var(--txt2);
  border-radius:4px;padding:2px 7px;font-size:.68rem;cursor:pointer;transition:all .15s;}
.preset-btn:hover{background:var(--blue);color:#fff;border-color:var(--blue);}
.price-filter-group{display:flex;gap:4px;}
.price-btn{background:var(--bg2);border:1px solid var(--bdr);color:var(--txt2);
  border-radius:4px;padding:3px 10px;font-size:.73rem;cursor:pointer;transition:all .15s;}
.price-btn.active{background:var(--blue);color:#fff;border-color:var(--blue);}
.btn-reset{background:var(--bg2);border:1px solid var(--bdr);color:var(--txt2);
  border-radius:5px;padding:5px 14px;font-size:.74rem;cursor:pointer;width:100%;transition:all .15s;}
.btn-reset:hover{border-color:var(--txt2);color:var(--txt);}

/* TABLE CARD */
.tbl-card{background:var(--bg1);border:1px solid var(--bdr);border-radius:10px;overflow:hidden;}
.tbl-hdr{background:linear-gradient(90deg,#1a3a5c,#162235);padding:.7rem 1.25rem;
  display:flex;align-items:center;justify-content:space-between;}
.tbl-hdr-title{color:#fff;font-weight:600;font-size:.88rem;}
.tbl-hdr-sub{color:var(--txt2);font-size:.72rem;}

/* DATATABLES */
table.dataTable{color:var(--txt)!important;font-size:.78rem;}
table.dataTable thead th{background:#152030!important;color:#7ab5e0!important;
  border-color:var(--bdr)!important;white-space:nowrap;font-size:.75rem;padding:7px 10px!important;}
table.dataTable tbody tr{background:var(--bg1)!important;}
table.dataTable tbody tr:nth-child(even){background:#111820!important;}
table.dataTable tbody tr:hover{background:#192840!important;}
table.dataTable tbody td{border-color:#1e2530!important;vertical-align:middle;padding:5px 8px!important;}
.dataTables_wrapper .dataTables_filter input,
.dataTables_wrapper .dataTables_length select{background:var(--bg0);border:1px solid var(--bdr);
  color:var(--txt);border-radius:6px;padding:4px 10px;}
.dataTables_wrapper .dataTables_info,
.dataTables_wrapper .dataTables_paginate{color:var(--txt2);font-size:.74rem;}
.dataTables_wrapper .paginate_button{color:var(--txt2)!important;border-radius:5px!important;}
.dataTables_wrapper .paginate_button.current{background:var(--blue)!important;border-color:var(--blue)!important;color:#fff!important;}
.dataTables_wrapper .paginate_button:hover:not(.disabled){background:var(--blue2)!important;border-color:var(--blue2)!important;color:#fff!important;}
.dt-buttons .btn{background:#1c2a3a!important;border-color:var(--bdr)!important;color:var(--txt)!important;font-size:.75rem;margin-right:4px;}
.dt-buttons .btn:hover{background:#253a50!important;}

/* BADGES */
.badge-gain{background:#0d2b1a;color:#3fb950;padding:2px 6px;border-radius:4px;font-weight:600;font-size:.74rem;}
.badge-dd{background:#2b0d0d;color:#f85149;padding:2px 6px;border-radius:4px;font-weight:600;font-size:.74rem;}
.badge-win{background:#0d1f2b;color:#58a6ff;padding:2px 6px;border-radius:4px;font-weight:600;font-size:.74rem;}
.badge-pf{background:#1a1f0d;color:#a3d977;padding:2px 6px;border-radius:4px;font-weight:600;font-size:.74rem;}
.lnk{color:#58a6ff;font-weight:600;text-decoration:none;}
.lnk:hover{text-decoration:underline;}
.fav-star{cursor:pointer;font-size:1rem;color:#30363d;transition:color .15s,transform .15s;user-select:none;display:inline-block;line-height:1;}
.fav-star:hover{color:#d29922;transform:scale(1.2);}
.fav-star.fav-on{color:#f0c040;text-shadow:0 0 6px #f0c04066;}

/* NOUISLIDER OVERRIDES */
.noUi-target{background:var(--bg0);border:1px solid var(--bdr);border-radius:4px;box-shadow:none;}
.noUi-connect{background:var(--blue2);}
.noUi-handle{background:#2a6099;border:1px solid #3a80bb;border-radius:3px;box-shadow:none;width:14px!important;height:14px!important;right:-7px!important;top:-5px!important;cursor:ew-resize;}
.noUi-handle::before,.noUi-handle::after{display:none;}
.noUi-horizontal{height:4px;}
.range-labels{display:flex;justify-content:space-between;font-size:.7rem;color:var(--txt2);margin-top:4px;}
.range-labels span{color:#58a6ff;font-weight:600;}

/* PROGRESS OVERLAY */
.prog-overlay{display:none;position:fixed;inset:0;background:rgba(0,0,0,.75);z-index:9999;
  align-items:center;justify-content:center;backdrop-filter:blur(3px);}
.prog-box{background:var(--bg1);border:1px solid var(--bdr);border-radius:14px;
  padding:2rem 2.5rem;min-width:360px;text-align:center;}
.prog-title{font-size:1rem;font-weight:600;margin-bottom:1rem;}
.prog-track{background:var(--bg0);border-radius:6px;height:8px;margin:1rem 0;overflow:hidden;}
.prog-fill{background:linear-gradient(90deg,var(--blue),#58a6ff);height:8px;border-radius:6px;transition:width .4s ease;}
.prog-msg{color:var(--txt2);font-size:.82rem;}
.prog-pct{font-size:1.4rem;font-weight:700;color:#58a6ff;margin-top:.5rem;}
</style>
</head>
<body>

<div class="topbar">
  <div style="display:flex;align-items:center;">
    <h1>📊 MT5 Signals Screener</h1>
    <span class="topbar-badge">MQL5 · MT5</span>
  </div>
  <div class="topbar-right">
    <span class="topbar-meta">Updated: %%SCRAPED_AT%% &nbsp;·&nbsp; <span id="totalCount">%%TOTAL%%</span> signals</span>
    <button class="btn-update" id="updateBtn" onclick="triggerUpdate()">🔄 Update Data</button>
  </div>
</div>

<div class="main-wrap">

  <div class="section-lbl">⭐ Top Rated Signals <small style="font-size:.65rem;font-weight:400;letter-spacing:0;text-transform:none;color:var(--txt2)">&nbsp;composite score: win rate · profit factor · risk/reward · longevity · popularity</small></div>
  <div class="row g-3 mb-3">
    <div class="col-12 col-xl-8">
      <div id="topPicksWrap" class="picks-wrap" style="margin-bottom:0"></div>
    </div>
    <div class="col-12 col-xl-4">
      <div class="chart-card" style="min-height:260px">
        <div class="chart-hdr">Radar Comparison &nbsp;<small id="radarLbl" style="font-weight:400;color:var(--txt2)">Top 5 — click legend name to open signal · click card to highlight</small></div>
        <canvas id="radarChart" style="max-height:300px"></canvas>
      </div>
    </div>
  </div>

  <div class="section-lbl">📈 Overview</div>
  <div id="statsRow" class="row g-3 mb-3"></div>

  <div class="row g-3 mb-3">
    <div class="col-12">
      <div class="chart-card">
        <div class="chart-hdr">Gain % vs Drawdown % &nbsp;<small style="font-weight:400">(bubble size = subscribers · color = win rate · updates with filters)</small></div>
        <canvas id="scatterChart" style="max-height:340px"></canvas>
      </div>
    </div>
  </div>

  <div class="section-lbl">🔍 Filters</div>
  <div class="filter-box">
    <div class="row g-3 align-items-start">

      <!-- Gain range (log scale) -->
      <div class="col-12 col-md-6 col-xl-3">
        <label>Gain %</label>
        <div id="gainRange" style="margin:10px 4px 4px"></div>
        <div class="range-labels"><span id="gainLo">0</span><span id="gainHi">∞</span></div>
        <div class="preset-btns mt-1">
          <button class="preset-btn" onclick="setGainRange(0,100)">Any</button>
          <button class="preset-btn" onclick="setGainRange(100,999999)">100%+</button>
          <button class="preset-btn" onclick="setGainRange(1000,999999)">1K%+</button>
          <button class="preset-btn" onclick="setGainRange(10000,999999)">10K%+</button>
          <button class="preset-btn" onclick="setGainRange(0,1000)">0–1K%</button>
          <button class="preset-btn" onclick="setGainRange(1000,10000)">1K–10K%</button>
        </div>
      </div>

      <!-- Drawdown range -->
      <div class="col-6 col-md-3 col-xl-2">
        <label>Drawdown %</label>
        <div id="ddRange" style="margin:10px 4px 4px"></div>
        <div class="range-labels"><span id="ddLo">0</span><span id="ddHi">100</span></div>
        <div class="preset-btns mt-1">
          <button class="preset-btn" onclick="setRange('dd',0,100)">Any</button>
          <button class="preset-btn" onclick="setRange('dd',0,20)">≤20%</button>
          <button class="preset-btn" onclick="setRange('dd',0,40)">≤40%</button>
          <button class="preset-btn" onclick="setRange('dd',40,100)">High DD</button>
        </div>
      </div>

      <!-- Win Rate range -->
      <div class="col-6 col-md-3 col-xl-2">
        <label>Win Rate %</label>
        <div id="winRange" style="margin:10px 4px 4px"></div>
        <div class="range-labels"><span id="winLo">0</span><span id="winHi">100</span></div>
        <div class="preset-btns mt-1">
          <button class="preset-btn" onclick="setRange('win',0,100)">Any</button>
          <button class="preset-btn" onclick="setRange('win',50,100)">50%+</button>
          <button class="preset-btn" onclick="setRange('win',60,100)">60%+</button>
          <button class="preset-btn" onclick="setRange('win',70,100)">70%+</button>
        </div>
      </div>

      <!-- Profit Factor range -->
      <div class="col-6 col-md-3 col-xl-2">
        <label>Profit Factor</label>
        <div id="pfRange" style="margin:10px 4px 4px"></div>
        <div class="range-labels"><span id="pfLo">0.0</span><span id="pfHi">10.0</span></div>
        <div class="preset-btns mt-1">
          <button class="preset-btn" onclick="setRange('pf',0,10)">Any</button>
          <button class="preset-btn" onclick="setRange('pf',1,10)">1.0+</button>
          <button class="preset-btn" onclick="setRange('pf',1.5,10)">1.5+</button>
          <button class="preset-btn" onclick="setRange('pf',2,10)">2.0+</button>
        </div>
      </div>

      <!-- Weeks range -->
      <div class="col-6 col-md-3 col-xl-2">
        <label>Weeks Active</label>
        <div id="wkRange" style="margin:10px 4px 4px"></div>
        <div class="range-labels"><span id="wkLo">0</span><span id="wkHi">500</span></div>
        <div class="preset-btns mt-1">
          <button class="preset-btn" onclick="setRange('wk',0,500)">Any</button>
          <button class="preset-btn" onclick="setRange('wk',12,500)">3mo+</button>
          <button class="preset-btn" onclick="setRange('wk',26,500)">6mo+</button>
          <button class="preset-btn" onclick="setRange('wk',52,500)">1yr+</button>
        </div>
      </div>

      <!-- Trades range -->
      <div class="col-6 col-md-3 col-xl-2">
        <label>Trades</label>
        <div id="trRange" style="margin:10px 4px 4px"></div>
        <div class="range-labels"><span id="trLo">0</span><span id="trHi">5000</span></div>
        <div class="preset-btns mt-1">
          <button class="preset-btn" onclick="setRange('tr',0,5000)">Any</button>
          <button class="preset-btn" onclick="setRange('tr',50,5000)">50+</button>
          <button class="preset-btn" onclick="setRange('tr',200,5000)">200+</button>
          <button class="preset-btn" onclick="setRange('tr',500,5000)">500+</button>
        </div>
      </div>

      <!-- Subscribers range -->
      <div class="col-6 col-md-3 col-xl-2">
        <label>Subscribers</label>
        <div id="subRange" style="margin:10px 4px 4px"></div>
        <div class="range-labels"><span id="subLo">0</span><span id="subHi">500</span></div>
        <div class="preset-btns mt-1">
          <button class="preset-btn" onclick="setRange('sub',0,500)">Any</button>
          <button class="preset-btn" onclick="setRange('sub',10,500)">10+</button>
          <button class="preset-btn" onclick="setRange('sub',50,500)">50+</button>
          <button class="preset-btn" onclick="setRange('sub',100,500)">100+</button>
        </div>
      </div>

      <!-- Price + Favorites + Reset -->
      <div class="col-12 col-md-6 col-xl-3">
        <div class="row g-2">
          <div class="col-12">
            <label>Price</label>
            <div class="price-filter-group">
              <button class="price-btn active" onclick="setPriceFilter('all',this)">All</button>
              <button class="price-btn" onclick="setPriceFilter('free',this)">Free</button>
              <button class="price-btn" onclick="setPriceFilter('paid',this)">Paid</button>
            </div>
          </div>
          <div class="col-12">
            <label>Favorites</label>
            <div class="price-filter-group">
              <button class="price-btn" id="favFilterBtn" onclick="toggleFavFilter(this)">⭐ Show Favorites</button>
              <button class="price-btn" onclick="clearAllFavs()" title="Clear all" style="padding:3px 8px">✕</button>
            </div>
          </div>
          <div class="col-12">
            <button class="btn-reset" onclick="resetFilters()">↺ Reset All Filters</button>
          </div>
        </div>
      </div>

    </div>
  </div>

  <div class="tbl-card">
    <div class="tbl-hdr">
      <span class="tbl-hdr-title">All Signals</span>
      <span class="tbl-hdr-sub">Sorted across all pages · click headers to sort · numeric sort applied</span>
    </div>
    <div class="p-3">
      <table id="tbl" class="table table-sm table-bordered w-100"></table>
    </div>
  </div>

</div>

<div id="progOverlay" class="prog-overlay">
  <div class="prog-box">
    <div class="prog-title">🔄 Fetching Latest Data from MQL5</div>
    <div class="prog-track"><div id="progFill" class="prog-fill" style="width:0%"></div></div>
    <div id="progMsg" class="prog-msg">Starting…</div>
    <div id="progPct" class="prog-pct">0%</div>
  </div>
</div>

<script src="https://code.jquery.com/jquery-3.7.1.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/bootstrap@5.3.2/dist/js/bootstrap.bundle.min.js"></script>
<script src="https://cdn.datatables.net/1.13.8/js/jquery.dataTables.min.js"></script>
<script src="https://cdn.datatables.net/1.13.8/js/dataTables.bootstrap5.min.js"></script>
<script src="https://cdn.datatables.net/buttons/2.4.2/js/dataTables.buttons.min.js"></script>
<script src="https://cdn.datatables.net/buttons/2.4.2/js/buttons.bootstrap5.min.js"></script>
<script src="https://cdnjs.cloudflare.com/ajax/libs/jszip/3.10.1/jszip.min.js"></script>
<script src="https://cdn.datatables.net/buttons/2.4.2/js/buttons.html5.min.js"></script>
<script src="https://cdn.datatables.net/buttons/2.4.2/js/buttons.print.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/chart.js@4.4.0/dist/chart.umd.min.js"></script>
<script src="https://cdn.jsdelivr.net/npm/nouislider@15.8.1/dist/nouislider.min.js"></script>
<script>
const RAW  = %%ROWS_JSON%%;
const KEYS = %%COLS_JSON%%;

// ── Utilities ─────────────────────────────────────────────────────────────────
function parseNum(s){return parseFloat((s||"0").replace(/[,\s%]/g,""))||0;}
function fmtK(n){return n>=1e6?`${(n/1e6).toFixed(1)}M`:n>=1e3?`${(n/1e3).toFixed(0)}K`:Math.round(n).toString();}

// ── Scoring ───────────────────────────────────────────────────────────────────
const COLORS=['#58a6ff','#3fb950','#f0a500','#f85149','#bc8cff','#39d353'];

function scoreSignal(r){
  const gain=parseNum(r["Gain %"]),dd=parseNum(r["Drawdown %"]);
  const win=parseNum(r["Win %"]),pf=parseNum(r["Profit Factor"]);
  const wks=parseNum(r["Weeks"]),subs=parseNum(r["Subscribers"]);
  const trades=parseNum(r["Trades"]);
  if(win<50||dd>60||pf<1.0||wks<12||trades<50||gain<=0) return -1;
  const sWin  =Math.max(0,(win-50)/50);
  const sPF   =Math.min((pf-1)/4,1);
  const calmar=Math.min(gain/Math.max(dd,1),50)/50;
  const sLong =Math.min(wks/156,1);
  const sPop  =Math.min(Math.log10(subs+1)/3,1);
  return sWin*0.30+sPF*0.25+calmar*0.25+sLong*0.15+sPop*0.05;
}

function scoreAxes(r){
  const gain=parseNum(r["Gain %"]),dd=parseNum(r["Drawdown %"]);
  const win=parseNum(r["Win %"]),pf=parseNum(r["Profit Factor"]);
  const wks=parseNum(r["Weeks"]);
  return [Math.min(win,100),Math.min(pf/5*100,100),
          Math.min(gain/Math.max(dd,1)/50*100,100),Math.min(wks/156*100,100),Math.max(0,100-dd)];
}

function scoreColor(s){const p=s*100;return p>=75?'#3fb950':p>=60?'#a3d977':p>=45?'#d29922':'#f85149';}

function pickBadges(r){
  const b=[];
  if(parseNum(r["Win %"])>=70)          b.push("🏆 Win Rate");
  if(parseNum(r["Profit Factor"])>=2.5) b.push("💎 Strong PF");
  if(parseNum(r["Weeks"])>=104)          b.push("⏰ 2yr+ Track");
  if(parseNum(r["Drawdown %"])<=20)      b.push("🛡 Low DD");
  if(parseNum(r["Subscribers"])>=100)    b.push("👥 Popular");
  return b.slice(0,3);
}

// ── Stats ─────────────────────────────────────────────────────────────────────
function numsOf(k){return RAW.map(r=>parseNum(r[k])).filter(v=>v>0);}
function avg(a){return a.length?(a.reduce((s,v)=>s+v,0)/a.length):0;}
function median(a){const s=[...a].sort((x,y)=>x-y);const m=Math.floor(s.length/2);return s.length%2?s[m]:(s[m-1]+s[m])/2;}
function pctile(a,p){const s=[...a].sort((x,y)=>x-y);return s[Math.floor(s.length*p/100)]||0;}

const gains=numsOf("Gain %"),dds=numsOf("Drawdown %"),wins=numsOf("Win %");
const qualifying=RAW.filter(r=>scoreSignal(r)>0).length;

const STAT_META=[
  ["Total Signals",   RAW.length,                          "#58a6ff","sv-total"],
  ["Median Gain",     fmtK(median(gains))+"%",             "#3fb950","sv-median"],
  ["95th Pct Gain",   fmtK(pctile(gains,95))+"%",         "#a3d977","sv-p95"],
  ["Avg Drawdown",    avg(dds).toFixed(1)+"%",             "#f85149","sv-dd"],
  ["Avg Win Rate",    avg(wins).toFixed(1)+"%",            "#d29922","sv-win"],
  ["Quality Signals", qualifying,                           "#bc8cff","sv-quality"],
];
STAT_META.forEach(([lbl,val,color,id])=>{
  document.getElementById("statsRow").innerHTML+=
    `<div class="col-6 col-md-4 col-xl-2"><div class="stat-card">
       <div class="stat-val" id="${id}" style="color:${color}">${val}</div>
       <div class="stat-lbl">${lbl}</div>
     </div></div>`;
});

// ── Top Picks ─────────────────────────────────────────────────────────────────
const scored=RAW.map(r=>({r,s:scoreSignal(r)})).filter(x=>x.s>0).sort((a,b)=>b.s-a.s).slice(0,10);
const TOP5=scored.slice(0,5);

scored.forEach(({r,s},i)=>{
  const pct=Math.round(s*100),col=scoreColor(s);
  const badges=pickBadges(r).map(b=>`<span class="pick-badge">${b}</span>`).join("");
  document.getElementById("topPicksWrap").innerHTML+=
    `<div class="pick-card" id="pickCard${i}" onclick="highlightRadar(${i})">
       <div class="pick-rank-row">
         <span class="pick-rank">#${i+1}</span>
         <span class="pick-score" style="background:${col}22;color:${col}">${pct}/100</span>
       </div>
       <a class="pick-name" href="${r.__link||'#'}" target="_blank" title="${r.Name||''}">${r.Name||'—'}</a>
       <div class="pick-bar-track"><div class="pick-bar-fill" style="width:${pct}%;background:${col}"></div></div>
       <div class="pick-stats">
         <span class="pick-stat-lbl">Gain</span><span class="pick-stat-val" style="color:#3fb950">${r["Gain %"]||'—'}</span>
         <span class="pick-stat-lbl">DD</span><span class="pick-stat-val" style="color:#f85149">${r["Drawdown %"]||'—'}</span>
         <span class="pick-stat-lbl">Win%</span><span class="pick-stat-val">${r["Win %"]||'—'}</span>
         <span class="pick-stat-lbl">PF</span><span class="pick-stat-val">${r["Profit Factor"]||'—'}</span>
         <span class="pick-stat-lbl">Weeks</span><span class="pick-stat-val">${r["Weeks"]||'—'}</span>
         <span class="pick-stat-lbl">Subs</span><span class="pick-stat-val">${r["Subscribers"]||'—'}</span>
       </div>
       <div class="pick-badges">${badges}</div>
     </div>`;
});

// ── Scatter Chart ─────────────────────────────────────────────────────────────
function rowsToScatterPts(rows){
  return rows.filter(r=>parseNum(r["Gain %"])>0&&parseNum(r["Drawdown %"])>0).map(r=>({
    x:Math.log10(parseNum(r["Gain %"])+1)*100,
    y:parseNum(r["Drawdown %"]),
    r:Math.max(3,Math.min(16,Math.sqrt(parseNum(r["Subscribers"])+1)*1.5)),
    _n:r.Name,_g:r["Gain %"],_d:r["Drawdown %"],_w:r["Win %"],
  }));
}
function ptsToColors(pts){
  return pts.map(p=>{const w=parseNum(p._w),t=Math.min(w/100,1);
    return `rgba(${Math.round(248-190*t)},${Math.round(13+156*t)},${Math.round(13+242*t)},0.65)`;
  });
}
const initPts=rowsToScatterPts(RAW);
const scatterChart=new Chart(document.getElementById("scatterChart"),{
  type:"bubble",
  data:{datasets:[{data:initPts,pointBackgroundColor:ptsToColors(initPts),pointBorderColor:"transparent",pointBorderWidth:0}]},
  options:{
    responsive:true,maintainAspectRatio:false,animation:{duration:0},
    plugins:{legend:{display:false},tooltip:{callbacks:{label:c=>{
      const d=c.raw;return [`${d._n||''}`,`Gain: ${d._g}  DD: ${d._d}  Win: ${d._w}`];
    }}}},
    scales:{
      x:{title:{display:true,text:"Gain % (log scale)",color:"#8b949e"},grid:{color:"#1e2530"},
         ticks:{color:"#8b949e",callback:v=>{const g=Math.round(Math.pow(10,v/100)-1);return g>=1000?`${Math.round(g/1000)}K%`:`${g}%`;}}},
      y:{title:{display:true,text:"Drawdown %",color:"#8b949e"},grid:{color:"#1e2530"},
         ticks:{color:"#8b949e",callback:v=>`${v}%`}},
    }
  }
});

// ── Radar Chart ───────────────────────────────────────────────────────────────
let radarChart=null,activePickIdx=null;
function buildRadar(items){
  if(radarChart) radarChart.destroy();
  radarChart=new Chart(document.getElementById("radarChart"),{
    type:"radar",
    data:{
      labels:["Win Rate","Profit Factor","Risk/Reward","Track Record","Safety"],
      datasets:items.map(({r},i)=>({
        label:(r.Name||'').length>22?(r.Name||'').slice(0,21)+'…':(r.Name||'—'),
        data:scoreAxes(r),
        borderColor:COLORS[i%COLORS.length],
        backgroundColor:COLORS[i%COLORS.length]+"25",
        borderWidth:2,pointRadius:3,pointBackgroundColor:COLORS[i%COLORS.length],
      }))
    },
    options:{
      responsive:true,maintainAspectRatio:false,animation:{duration:300},
      plugins:{
        legend:{
          labels:{color:"#c9d1d9",font:{size:10},boxWidth:12,
            generateLabels:(chart)=>Chart.defaults.plugins.legend.labels.generateLabels(chart).map(l=>({...l,fontStyle:'underline'}))},
          onClick:(e,legendItem)=>{
            const sig=items[legendItem.datasetIndex];
            if(sig&&sig.r.__link) window.open(sig.r.__link,'_blank');
          }
        }
      },
      scales:{r:{min:0,max:100,
        ticks:{color:"#8b949e",backdropColor:"transparent",stepSize:25,font:{size:9}},
        grid:{color:"#2a3540"},angleLines:{color:"#2a3540"},
        pointLabels:{color:"#c9d1d9",font:{size:10}},
      }}
    }
  });
}
buildRadar(TOP5);

function highlightRadar(i){
  document.querySelectorAll(".pick-card").forEach(c=>c.classList.remove("active"));
  if(activePickIdx===i){activePickIdx=null;buildRadar(TOP5);document.getElementById("radarLbl").textContent="Top 5 signals — click a card to compare";}
  else{
    activePickIdx=i;
    document.getElementById("pickCard"+i).classList.add("active");
    const sel=scored[i];
    const others=TOP5.filter(x=>x!==sel).slice(0,4);
    buildRadar([sel,...others]);
    document.getElementById("radarLbl").textContent=`Highlighting: ${(sel.r.Name||'').slice(0,25)}`;
  }
}

// ── Favorites ─────────────────────────────────────────────────────────────────
let favorites=new Set(JSON.parse(localStorage.getItem('mt5_fav')||'[]'));
let showFavOnly=false;

function saveFavs(){ localStorage.setItem('mt5_fav',JSON.stringify([...favorites])); }

function toggleFav(id,e){
  if(e){e.stopPropagation();e.preventDefault();}
  if(favorites.has(id)) favorites.delete(id); else favorites.add(id);
  saveFavs();
  table.draw(false);
  updateFavBtn();
}

function clearAllFavs(){
  if(!favorites.size) return;
  if(!confirm(`Remove all ${favorites.size} favorites?`)) return;
  favorites.clear(); saveFavs();
  table.draw(false); updateFavBtn();
}

function updateFavBtn(){
  const btn=document.getElementById('favFilterBtn');
  if(btn) btn.textContent=`⭐ Show Favorites${favorites.size?' ('+favorites.size+')':''}`;
  if(showFavOnly&&!favorites.size){ showFavOnly=false; btn&&btn.classList.remove('active'); applyFilters(); }
}

function toggleFavFilter(btn){
  showFavOnly=!showFavOnly;
  btn.classList.toggle('active',showFavOnly);
  applyFilters();
}

updateFavBtn();

// ── DataTable ─────────────────────────────────────────────────────────────────
const NUM_SET=new Set(["Gain %","Drawdown %","Win %","Profit Factor","Expected Payoff",
                        "Weeks","Trades","Subscribers","Subscriber Funds","Balance","Rank"]);
function makeCol(key){
  const c={title:key,data:key,defaultContent:""};
  if(key==="Name"){
    c.render=(d,t,r)=>t==="display"&&r.__link?`<a class="lnk" href="${r.__link}" target="_blank">${d}</a>`:d;
    c.width="180px";
  } else if(NUM_SET.has(key)){
    const badge=key==="Gain %"?"badge-gain":key==="Drawdown %"?"badge-dd":key==="Win %"?"badge-win":key==="Profit Factor"?"badge-pf":null;
    c.render=(d,t)=>{
      if(t==="sort"||t==="type") return parseNum(d);
      return badge?`<span class="${badge}">${d}</span>`:d;
    };
  }
  return c;
}

const starCol={
  title:'★',data:'__id',orderable:false,searchable:false,width:'32px',className:'text-center',
  render:(id,t)=>{
    if(!id||t!=='display') return '';
    return `<span class="fav-star${favorites.has(id)?' fav-on':''}" onclick="toggleFav('${id}',event)" title="${favorites.has(id)?'Remove from':'Add to'} favorites">★</span>`;
  },
};

const columns=[starCol,...KEYS.map(makeCol)];
columns.push({data:"__link",visible:false,searchable:false});
columns.push({data:"__id",visible:false,searchable:false});

const gainIdx=KEYS.indexOf("Gain %")+1; // +1 for star column
const table=$('#tbl').DataTable({
  data:RAW,columns,
  pageLength:50,
  lengthMenu:[[25,50,100,250,-1],[25,50,100,250,"All"]],
  order:[[gainIdx>=0?gainIdx:0,"desc"]],
  dom:'<"row mb-2"<"col-sm-6"B><"col-sm-6"f>>rt<"row mt-2"<"col-sm-5"i><"col-sm-7"p>>',
  buttons:[
    {extend:"csvHtml5",  text:"⬇ CSV",   className:"btn btn-sm"},
    {extend:"excelHtml5",text:"⬇ Excel", className:"btn btn-sm"},
    {extend:"print",     text:"🖨 Print",  className:"btn btn-sm"},
  ],
  scrollX:true,
  language:{search:"🔍 ",searchPlaceholder:"Search signals…"},
});

// Update charts/stats when the DataTable search box filters rows
table.on('search.dt draw.dt', function(){ updateChartsAndStats(); });

// ── Dual range sliders (noUiSlider) ──────────────────────────────────────────
function s2g(v){return v<=0?0:Math.round(Math.pow(10,v*6/100)-1);}
function g2s(g){return g<=0?0:Math.min(100,Math.log10(g+1)/6*100);}
function fmtGain(v){const g=Math.round(s2g(v));return g>=999990?'∞':g>=1000?fmtK(g)+'%':g+'%';}

const F={gain:[0,100],dd:[0,100],win:[0,100],pf:[0,10],wk:[0,500],tr:[0,5000],sub:[0,500]};

function mkSlider(id,key,cfg){
  const el=document.getElementById(id);
  noUiSlider.create(el,{start:cfg.start||[cfg.min,cfg.max],connect:true,
    range:{min:cfg.min,max:cfg.max},step:cfg.step||1});
  el.noUiSlider.on('update',(v)=>{
    F[key]=[+v[0],+v[1]];
    const loEl=document.getElementById(key+'Lo'), hiEl=document.getElementById(key+'Hi');
    if(loEl) loEl.textContent=cfg.fmt?cfg.fmt(+v[0],true):Math.round(+v[0]);
    if(hiEl) hiEl.textContent=cfg.fmt?cfg.fmt(+v[1],false):(+v[1]>=cfg.max?(cfg.cap||Math.round(+v[1])):Math.round(+v[1]));
  });
  el.noUiSlider.on('change',()=>applyFilters());
}

mkSlider('gainRange','gain',{min:0,max:100,start:[0,100],step:.5,fmt:(v)=>fmtGain(v)});
mkSlider('ddRange','dd',  {min:0,max:100,start:[0,100], fmt:(v)=>v+'%', cap:'100%'});
mkSlider('winRange','win',{min:0,max:100,start:[0,100], fmt:(v)=>v+'%', cap:'100%'});
mkSlider('pfRange','pf',  {min:0,max:10, start:[0,10],  step:.1,fmt:(v)=>(+v).toFixed(1),cap:'∞'});
mkSlider('wkRange','wk',  {min:0,max:500,start:[0,500], cap:'500+'});
mkSlider('trRange','tr',  {min:0,max:5000,start:[0,5000],step:50,cap:'5K+'});
mkSlider('subRange','sub',{min:0,max:500,start:[0,500], cap:'500+'});

function setGainRange(lo,hi){
  document.getElementById('gainRange').noUiSlider.set([g2s(lo),g2s(Math.min(hi,999999))]);
  applyFilters();
}
function setRange(key,lo,hi){
  document.getElementById(key+'Range').noUiSlider.set([lo,hi]);
  applyFilters();
}

let priceFilter='all';
function setPriceFilter(mode,btn){
  priceFilter=mode;
  document.querySelectorAll(".price-btn").forEach(b=>b.classList.remove("active"));
  btn.classList.add("active");
  applyFilters();
}

// ── Live stats + chart update ─────────────────────────────────────────────────
function updateChartsAndStats(){
  const filtered=[];
  table.rows({filter:"applied"}).data().each(r=>filtered.push(r));
  const fg=filtered.map(r=>parseNum(r["Gain %"])).filter(v=>v>0);
  const fd=filtered.map(r=>parseNum(r["Drawdown %"])).filter(v=>v>0);
  const fw=filtered.map(r=>parseNum(r["Win %"])).filter(v=>v>0);
  const fq=filtered.filter(r=>scoreSignal(r)>0).length;
  document.getElementById("sv-total").textContent   =filtered.length;
  document.getElementById("sv-median").textContent  =fg.length?fmtK(median(fg))+"%":"—";
  document.getElementById("sv-p95").textContent     =fg.length?fmtK(pctile(fg,95))+"%":"—";
  document.getElementById("sv-dd").textContent      =fd.length?avg(fd).toFixed(1)+"%":"—";
  document.getElementById("sv-win").textContent     =fw.length?avg(fw).toFixed(1)+"%":"—";
  document.getElementById("sv-quality").textContent =fq;
  const pts=rowsToScatterPts(filtered);
  scatterChart.data.datasets[0].data=pts;
  scatterChart.data.datasets[0].pointBackgroundColor=ptsToColors(pts);
  scatterChart.update("none");
}

// ── Filter logic ──────────────────────────────────────────────────────────────
function applyFilters(){
  const gainMin=s2g(F.gain[0]),gainMaxEff=F.gain[1]>=99.9?Infinity:s2g(F.gain[1]);
  const ddLo=F.dd[0],ddHi=F.dd[1];
  const winLo=F.win[0],winHi=F.win[1];
  const pfLo=F.pf[0],pfHi=F.pf[1]>=9.9?Infinity:F.pf[1];
  const wkLo=F.wk[0],wkHi=F.wk[1]>=499?Infinity:F.wk[1];
  const trLo=F.tr[0],trHi=F.tr[1]>=4990?Infinity:F.tr[1];
  const subLo=F.sub[0],subHi=F.sub[1]>=499?Infinity:F.sub[1];

  $.fn.dataTable.ext.search=[];
  $.fn.dataTable.ext.search.push((_,__,___,row)=>{
    if(showFavOnly&&!favorites.has(row.__id))          return false;
    const gain=parseNum(row["Gain %"]);
    if(gain<gainMin||gain>gainMaxEff)                  return false;
    const dd=parseNum(row["Drawdown %"]);
    if(dd<ddLo||dd>ddHi)                              return false;
    const win=parseNum(row["Win %"]);
    if(win<winLo||win>winHi)                           return false;
    const pf=parseNum(row["Profit Factor"]);
    if(pf<pfLo||pf>pfHi)                              return false;
    const wk=parseNum(row["Weeks"]);
    if(wk<wkLo||wk>wkHi)                              return false;
    const tr=parseNum(row["Trades"]);
    if(tr<trLo||tr>trHi)                              return false;
    const sub=parseNum(row["Subscribers"]);
    if(sub<subLo||sub>subHi)                           return false;
    const price=parseNum(row["Price (USD/mo)"]);
    if(priceFilter==="free"&&price>0)                  return false;
    if(priceFilter==="paid"&&price===0)                return false;
    return true;
  });
  table.draw();
  updateChartsAndStats();
}

function resetFilters(){
  ['gainRange','ddRange','winRange','wkRange','trRange','subRange'].forEach(id=>{
    const el=document.getElementById(id);
    const cfg={gainRange:[0,100],ddRange:[0,100],winRange:[0,100],
               wkRange:[0,500],trRange:[0,5000],subRange:[0,500]};
    el.noUiSlider.set(cfg[id]);
  });
  document.getElementById('pfRange').noUiSlider.set([0,10]);
  priceFilter="all";
  document.querySelectorAll(".price-btn").forEach((b,i)=>b.classList.toggle("active",i===0));
  showFavOnly=false;
  const fb=document.getElementById("favFilterBtn");
  if(fb) fb.classList.remove("active");
  $.fn.dataTable.ext.search=[];
  table.draw();
  updateChartsAndStats();
}

// ── Update button ─────────────────────────────────────────────────────────────
function triggerUpdate(){
  if(window.location.protocol==="file:"){
    alert("Live updates require the server.\n\nRun:  python server.py\nThen open:  http://localhost:5000");
    return;
  }
  const btn=document.getElementById("updateBtn");
  btn.disabled=true;btn.innerHTML="⏳ Updating…";
  document.getElementById("progOverlay").style.display="flex";
  fetch("/api/refresh",{method:"POST"})
    .then(r=>r.json()).then(()=>pollStatus())
    .catch(e=>{btn.disabled=false;btn.innerHTML="🔄 Update Data";
      document.getElementById("progOverlay").style.display="none";alert("Error: "+e);});
}
function pollStatus(){
  fetch("/api/status").then(r=>r.json()).then(s=>{
    const pct=s.total>0?Math.round(s.progress/s.total*100):0;
    document.getElementById("progFill").style.width=pct+"%";
    document.getElementById("progMsg").textContent=s.message||"";
    document.getElementById("progPct").textContent=pct+"%";
    if(s.status==="done"){setTimeout(()=>location.reload(),600);}
    else if(s.status==="error"){
      document.getElementById("progOverlay").style.display="none";
      document.getElementById("updateBtn").disabled=false;
      document.getElementById("updateBtn").innerHTML="🔄 Update Data";
      alert("Scrape error: "+s.message);
    }else{setTimeout(pollStatus,2000);}
  });
}
</script>
</body>
</html>"""

    html_out = (template
                .replace("%%ROWS_JSON%%", rows_json)
                .replace("%%COLS_JSON%%", cols_json)
                .replace("%%SCRAPED_AT%%", scraped_at)
                .replace("%%TOTAL%%", str(total)))

    with open(path, "w", encoding="utf-8") as f:
        f.write(html_out)
    print(f"  ✓ HTML   → {path}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    out_dir    = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(out_dir, exist_ok=True)
    scraped_at = datetime.now().strftime("%Y-%m-%d %H:%M")
    ts         = datetime.now().strftime("%Y%m%d_%H%M")

    print("=" * 64)
    print(f"  MT5 Signals Screener  |  {scraped_at}")
    print("=" * 64)

    signals = scrape_all()

    print("=" * 64)
    print(f"  Total signals collected: {len(signals)}")
    print("=" * 64)

    if not signals:
        print("\nNo data scraped — check your connection or MQL5 may be blocking requests.")
        sys.exit(1)

    df = pd.DataFrame(signals)
    print(f"\n  Generating reports...")

    excel_path = os.path.join(out_dir, f"mt5_signals_{ts}.xlsx")
    html_path  = os.path.join(out_dir, f"mt5_signals_{ts}.html")

    build_excel(df, excel_path)
    build_html(df, html_path, scraped_at)

    print(f"\n  Done!  Files are in:\n    {out_dir}")


if __name__ == "__main__":
    main()
